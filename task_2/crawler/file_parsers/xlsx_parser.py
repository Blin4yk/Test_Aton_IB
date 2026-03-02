from openpyxl import load_workbook
from .base import BaseParser


class XLSXParser(BaseParser):
    def parse(self, file_path: str) -> str:
        wb = load_workbook(file_path, data_only=True)
        sheets_text = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = ' '.join(str(cell) for cell in row if cell is not None)
                if row_text:
                    sheets_text.append(row_text)
        return '\n'.join(sheets_text)